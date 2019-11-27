import os
from mailmerge import MailMerge
from openpyxl import load_workbook


def write2docx(datum, template_name, output):
    """
    使用datum中的键值对来替换掉模板中的键，并生成名称为output的文件
    目前默认认为output的文件夹存在
    :param datum: dict键值对
    :param template_name: 模板的完整路径
    :param output: 输出的文件名
    :return:
    """
    doc = MailMerge(template_name)
    doc.merge(**datum)

    doc.write(output)


def write2one_docx(data, template, output, separator='page_break'):
    """
    根据data数组生成一个len(data)个的一个总文档
    :param data: 数组，键值对用来替换模板
    :param template: 模板完整名称
    :param output: 输出文件名称
    :param separator: 分隔符 默认为换页符
    :return:
    """
    doc = MailMerge(template)
    doc.merge_templates(data, separator=separator)

    doc.write(output)


def combine_word_documents(files, output):
    from docx import Document
    """
    合并多个文档为一个文档 目前会造成部分格式问题 需要保证输出文件所在的路径存在
    :param files: 数组，保存着要合并的文件路径
    :param output: 合并文件名称
    :return:
    """
    merged_document = Document()

    for index, file in enumerate(files):
        sub_doc = Document(file)

        # 在文档和文档之间添加一个换页
        if index < len(files) - 1:
            sub_doc.add_page_break()
        # 添加到总文件里
        for element in sub_doc.element.body:
            merged_document.element.body.append(element)

    merged_document.save(output)


def read_xlsx2dict(filename, sheet=None, callback=None):
    """
    读取excel文件，并得到[{}] 以第一行为键，其他行为指
    :param filename: 扩展名为xlsx的文件
    :param sheet: 要读取的sheet名称
    :param callback: callback(key, value)返回值用以代替原来的value
    :return: 数组
    """
    workbook = load_workbook(filename=filename)
    # 获取sheet
    sheet_names = workbook.sheetnames
    sheet = sheet_names[0] if not sheet else sheet
    table = workbook[sheet]
    # 把第一行作为键
    keys, index = None, 0
    # 按照行进行遍历
    for row in table.rows:
        line = []
        is_none = True
        # 获取每一个列
        for col in row:
            line.append(col.value)
            is_none = is_none and col.value is None
        # 第一列作为键
        if index == 0:
            keys = [key for key in line if key is not None]
        elif not is_none:
            datum = {}
            for i in range(len(keys)):
                if not callback:
                    datum[keys[i]] = callback(keys[i], line[i])
                else:
                    datum[keys[i]] = line[i]
            yield datum
        index += 1


def test_one_docx():
    datum = {'teacher_name': 'sky', 'day': '3', 'name': 'moon', 'date': '2019/11/26'}
    write2docx(datum, 'template.docx', 'one.docx')


def test_complete_docx():
    data = [
        {'teacher_name': 'sky', 'day': '3', 'name': 'moon', 'date': '2019/11/26'},
        {'teacher_name': '周志华', 'day': '3', 'name': 'moon', 'date': '2019/11/26'},
    ]
    write2one_docx(data, 'template.docx', 'two.docx', separator='textWrapping_break')


if __name__ == '__main__':
    test_complete_docx()
